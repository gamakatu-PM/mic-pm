# -*- coding: utf-8 -*-
"""총괄 단가 등록 + 현장 시트 값굳음 셀의 수식 복구 (성윤님 가격 확정 시 실행).

하는 일 (전부 B등급 — 바뀌기 전 값을 변경 이력에 남긴다):
  1. 「1.모듈총괄」에서 모듈명 정확 일치 행을 찾아 실행가격 기입 (주황/노랑 해제)
     행이 없으면 --after 앵커 다음에 새 행 삽입 (앵커 없으면 맨 끝)
  2. 전 현장 시트를 전수 조사 — 같은 모듈명의 단가 셀이 값으로 굳어 있으면
     총괄 INDEX/MATCH 수식으로 복구 (노랑 해제, 원값은 이력에 기록)
  3. 이 단가를 참조하는 시트 전부를 찾아 전파 범위를 보고 (다른 현장에 조용히
     번지는 것을 막기 위해 — 운영기준 §7.A)
  4. 「5.변경 이력」 기입, 버전 +1 저장

사용:
    python3 register_price.py 단가장_v40.xlsx \
        --set "퓨즈 1A=50" --after "퓨즈 1A=퓨즈 5A" --note "퓨즈 1A=SA0201B 앞단" \
        --set "CB 외함 400*600*90=120000"
"""
import argparse
import datetime
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.formula import ArrayFormula

sys.path.insert(0, str(Path(__file__).parent))
from common import (HISTORY_SHEET, MASTER_SHEET, find_master_names,
                    last_data_row, master_price_formula, set_cell)

TODAY = datetime.date.today().isoformat()
NOFILL = PatternFill(fill_type=None)


def fval(c):
    v = c.value
    return v.text if isinstance(v, ArrayFormula) else v


def find_price_col(ws):
    for row in ws.iter_rows(min_row=3, max_row=5):
        for c in row:
            if isinstance(c.value, str) and "실행단가" in c.value:
                return c.column
    return None


def parse_kv(pairs):
    out = {}
    for p in pairs or []:
        k, _, v = p.partition("=")
        out[k.strip()] = v.strip()
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--set", action="append", required=True,
                    help='"모듈명=가격" (반복 가능)')
    ap.add_argument("--after", action="append", help='"모듈명=앵커행 모듈명"')
    ap.add_argument("--note", action="append", help='"모듈명=비고 추가문"')
    ap.add_argument("-o", "--output")
    args = ap.parse_args()

    prices = {k: float(v) for k, v in parse_kv(args.set).items()}
    afters = parse_kv(args.after)
    notes = parse_kv(args.note)

    wb = openpyxl.load_workbook(args.workbook)
    names = find_master_names(wb)
    master = wb[MASTER_SHEET]
    history, report = [], {"master": [], "restored": [], "propagation": {}}

    for name, price in prices.items():
        if name in names:
            r = names[name]
            old = master.cell(row=r, column=4).value
            set_cell(master, f"D{r}", price, number_format="#,##0")
            for col in "ABCDE":
                master[f"{col}{r}"].fill = NOFILL
            if notes.get(name):
                cur = master.cell(row=r, column=5).value or ""
                set_cell(master, f"E{r}", (cur + " / " if cur else "") + notes[name])
            history.append(("B", f"{MASTER_SHEET} {r}행 「{name}」",
                            f"실행가격 {price:,.0f} 등록 (성윤님 확정 {TODAY}), 주황/노랑 해제",
                            f"가격 {'공란' if old in (None, '') else old}",
                            "가격 지우고 색 복원"))
            report["master"].append({"name": name, "row": r, "old": old, "new": price})
        else:
            anchor = afters.get(name)
            if anchor and anchor in names:
                r = names[anchor] + 1
                master.insert_rows(r)
                gubun = master.cell(row=names[anchor], column=1).value
            else:
                r = last_data_row(master) + 1
                gubun = "신규"
            set_cell(master, f"A{r}", gubun)
            set_cell(master, f"B{r}", name)
            set_cell(master, f"D{r}", price, number_format="#,##0")
            set_cell(master, f"E{r}", notes.get(name, f"성윤님 확정 {TODAY}"))
            history.append(("A", f"{MASTER_SHEET} {r}행",
                            f"「{name}」 신설, 실행가격 {price:,.0f} (성윤님 확정 {TODAY})",
                            "(없던 행)", "해당 행 삭제"))
            report["master"].append({"name": name, "row": r, "old": None, "new": price})
            names = find_master_names(wb)  # 삽입으로 행번호 밀림 반영

    # 현장 시트 전수: 값굳음 복구 + 전파 범위 수집
    for ws in wb.worksheets:
        a1 = ws["A1"].value
        if not (isinstance(a1, str) and a1.startswith("현장:")):
            continue
        pcol = find_price_col(ws)
        if pcol is None:
            continue
        for r in range(5, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if not isinstance(a, str) or a.strip() not in prices:
                continue
            nm = a.strip()
            report["propagation"].setdefault(nm, []).append(ws.title)
            pc = ws.cell(row=r, column=pcol)
            pv = fval(pc)
            if not (isinstance(pv, str) and pv.startswith("=")):
                old = pv
                set_cell(ws, f"{pc.column_letter}{r}", master_price_formula(f"A{r}"),
                         number_format="#,##0")
                pc.fill = NOFILL
                history.append(("B", f"{ws.title}!{pc.coordinate} 「{nm}」",
                                "값굳음 단가를 총괄 참조 수식으로 복구",
                                f"값 {old:,.0f}" if old is not None else "공란",
                                "수식 지우고 값 복원"))
                report["restored"].append(
                    {"sheet": ws.title, "cell": pc.coordinate, "name": nm, "old": old})

    hist = wb[HISTORY_SHEET]
    m = re.search(r"_v(\d+)", Path(args.workbook).name)
    version = (int(m.group(1)) if m else 0) + 1
    row = last_data_row(hist) + 1
    for grade, target, what, before, revert in history:
        for i, v in enumerate([f"v{version}", TODAY, grade, target, what, before, revert], 1):
            hist.cell(row=row, column=i, value=v)
        row += 1

    out = args.output or str(Path(args.workbook).with_name(f"CB모듈_단가장_v{version}.xlsx"))
    wb.save(out)
    report.update({"output": out, "version": version, "history_rows": len(history)})
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
