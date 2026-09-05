# -*- coding: utf-8 -*-
"""CB모듈_단가장 전수 검산. 결과 마지막 줄 `total_errors: 0`이어야 전달 가능.

검사 항목 (km-drawing-cost §4 검산 규칙의 코드화):
  1. 총괄 매칭 — 현장 시트 모듈명이 「1.모듈총괄」 B열과 정확히 일치하는가
     (불일치 = INDEX/MATCH 실패 = 금액 조용히 누락. ERROR)
  2. 수식 굳음 — 실행단가 셀이 총괄 참조 수식인가. 값으로 굳었으면 WARN
     (총괄 미등록 임시값일 수 있으므로 삭제하지 않고 알리기만 한다)
  3. 요율 — ▷ ×1.5 / ×1.8 행이 실행 합계를, 160% 행이 ×1.5 행을 참조하는가
     (160%를 실행×1.6으로 잘못 걸면 설계예산가가 통째로 틀어진다. ERROR)
  4. 재계산 — LibreOffice headless로 전체 재계산 후 #REF!·#N/A 등 오류값 0건
  5. 검산표 — 기구물 시트 「■ 검산」 판정에 '불일치' 0건 (재계산 값 기준)
  6. 객실 수 — ◆행 합계가 시트에 명시된 총 객실 수와 일치하는가

사용:  python3 validate.py 단가장.xlsx
"""
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

sys.path.insert(0, str(Path(__file__).parent))
from common import MASTER_SHEET, find_master_names

ERRVALS = ("#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!")
errors, warnings = [], []


def err(msg):
    errors.append(msg)
    print(f"  ERROR  {msg}")


def warn(msg):
    warnings.append(msg)
    print(f"  WARN   {msg}")


def fval(c):
    v = c.value
    return v.text if isinstance(v, ArrayFormula) else v


def site_sheets(wb):
    for ws in wb.worksheets:
        a1 = ws["A1"].value
        if isinstance(a1, str) and a1.startswith("현장:"):
            yield ws


def find_price_col(ws):
    for row in ws.iter_rows(min_row=3, max_row=5):
        for c in row:
            if isinstance(c.value, str) and "실행단가" in c.value:
                return c.column
    return None


def check_master_matching(wb):
    print("[1·2] 총괄 매칭 · 수식 굳음")
    names = find_master_names(wb)
    for ws in site_sheets(wb):
        pcol = find_price_col(ws)
        if pcol is None:
            warn(f"{ws.title}: 실행단가 열을 찾지 못함 — 구형 레이아웃일 수 있음")
            continue
        for r in range(6, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if not isinstance(a, str) or not a.strip():
                continue
            if a.startswith(("▣", "■", "◆", "▷", "★", "※", "합계", "품목", "모듈명")):
                continue
            pc = ws.cell(row=r, column=pcol)
            pv = fval(pc)
            name = a.strip() if not a.startswith("=") else None
            if isinstance(pv, str) and pv.startswith("="):
                if MASTER_SHEET in pv and name and name not in names:
                    err(f"{ws.title}!A{r} 「{name}」 총괄 미등록 — INDEX/MATCH 실패, 금액 누락됨")
            elif pv is not None and name:
                if name not in names:
                    warn(f"{ws.title}!{pc.coordinate} 「{name}」 단가가 값으로 굳음"
                         f"({pv:,.0f}) + 총괄 미등록 — 총괄 등록 필요")
                else:
                    warn(f"{ws.title}!{pc.coordinate} 「{name}」 단가가 값으로 굳음"
                         f"({pv:,.0f}) — 총괄 수식으로 복구 필요")


def check_rates(wb):
    """요율 세대 검사 (정본: 계약 ×1.5 / 견적 ×1.8 / 설계예산 ×2.1 — 2026-08-13 확정).

    구 160%(계약×1.6=실행×2.4)·구구 ×1.3/×1.4는 폐기 요율. 잔존 시트는 WARN으로 모아
    보고한다 (기존 현장 시트 일괄 정정은 성윤 차장 승인 후 별도 작업).
    신규 생성 시트는 요율 입력칸(노란 배수 칸) 참조 또는 ×2.1이어야 한다.
    """
    print("[3] 요율 세대 검사 (정본 ×1.5/×1.8/×2.1 — 구 160%·×1.3/×1.4 폐기)")
    legacy = []
    for ws in site_sheets(wb):
        gens = set()
        for row in ws.iter_rows(max_row=min(ws.max_row, 120)):
            a = row[0].value
            label = a if isinstance(a, str) else ""
            for c in row:
                v = fval(c)
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                vv = v.replace(" ", "")
                if "*1.3" in vv or "*1.4" in vv:
                    gens.add("폐기(×1.3/×1.4)")
                if "*1.6" in vv and ("▷" in label or "★" in label or "160%" in label):
                    gens.add("구160%(실행×2.4)")
                if "*2.1" in vv or re.search(r"\$[A-Z]+\$[5-9]\b", vv) and "TEXT(" in vv:
                    gens.add("신(×2.1)")
            if "160%" in label and not label.startswith("="):
                gens.add("구160%(실행×2.4)")
        bad = {g for g in gens if g.startswith(("구", "폐기"))}
        if bad:
            legacy.append(ws.title)
            warn(f"{ws.title}: 폐기 요율 잔존 {sorted(bad)} — 정본은 실행 ×2.1 "
                 f"(대외 제출 전 정정 필요)")
    print(f"  legacy_rate_sheets: {len(legacy)}")


def recalc(path):
    """LibreOffice headless 재계산 → 계산값이 캐시된 사본 경로 반환."""
    tmp = tempfile.mkdtemp(prefix="dgz_recalc_")
    subprocess.run(
        ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", tmp, str(path)],
        check=True, capture_output=True, timeout=300)
    out = Path(tmp) / (Path(path).stem + ".xlsx")
    return out


def check_computed(path):
    print("[4·5·6] 재계산 · 검산표 · 객실 수 (LibreOffice)")
    rp = recalc(path)
    wb = openpyxl.load_workbook(rp, data_only=True)
    nerr = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value in ERRVALS:
                    err(f"{ws.title}!{c.coordinate} 수식 오류값 {c.value}")
                    nerr += 1
                if isinstance(c.value, str) and c.value == "불일치":
                    err(f"{ws.title}!{c.coordinate} 검산표 불일치 — 수량이 계통도와 다름")
    # 객실 수: ◆행 합계 vs 헤더의 "(N실)" 표기
    for ws in wb.worksheets:
        a1 = ws["A1"].value
        if not (isinstance(a1, str) and a1.startswith("현장:")):
            continue
        declared = None
        for row in ws.iter_rows(min_row=1, max_row=5):
            for c in row:
                if isinstance(c.value, str):
                    m = re.search(r"\((\d+)\s*실\)", c.value)
                    if m:
                        declared = int(m.group(1))
        # 총합 열을 헤더에서 찾는다 — 행 오른쪽의 요율칸 숫자를 합계로 오인하지 않기 위해
        sum_col = None
        for c in ws[4]:
            h = c.value
            if isinstance(h, str) and ("총합" in h or h.startswith("객실 (")):
                sum_col = c.column
                break
        for r in range(4, 7):
            a = ws.cell(row=r, column=1).value
            if isinstance(a, str) and a.startswith("◆"):
                total = None
                if sum_col and isinstance(ws.cell(row=r, column=sum_col).value, (int, float)):
                    total = ws.cell(row=r, column=sum_col).value
                else:
                    for c in ws[r]:
                        if c.column > 2 and isinstance(c.value, (int, float)):
                            total = c.value  # 구형 시트 폴백: 마지막 숫자 = SUM 셀
                if declared and total and int(total) != declared:
                    err(f"{ws.title} ◆행 합계 {int(total)} ≠ 명시 객실 수 {declared}")
                elif declared and total:
                    print(f"  OK     {ws.title}: ◆행 합계 {int(total)} = 명시 {declared}")
    shutil.rmtree(rp.parent, ignore_errors=True)
    return wb


def report_totals(wb_computed):
    print("[요약] ★ 현장 총액 (재계산 값)")
    for ws in wb_computed.worksheets:
        a1 = ws["A1"].value
        if not (isinstance(a1, str) and a1.startswith("현장:")):
            continue
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if isinstance(a, str) and a.startswith("★"):
                vals = [c.value for c in ws[r]
                        if isinstance(c.value, (int, float)) and c.column > 2]
                if vals:
                    print(f"  {ws.title:20s} {a:32s} {vals[0]:>15,.0f}")


def main():
    path = sys.argv[1]
    wb = openpyxl.load_workbook(path)
    if MASTER_SHEET not in wb.sheetnames:
        err(f"「{MASTER_SHEET}」 시트가 없음 — 총괄은 파일에 언제나 존재해야 한다")
    else:
        check_master_matching(wb)
        check_rates(wb)
    wbc = check_computed(path)
    report_totals(wbc)
    print(f"\ntotal_warnings: {len(warnings)}")
    print(f"total_errors: {len(errors)}")
    sys.exit(1 if errors else 0)


if __name__ == "__main__":
    main()
