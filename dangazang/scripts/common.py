# -*- coding: utf-8 -*-
"""CB모듈_단가장 파이프라인 공통 모듈.

표준 시트 레이아웃(v40 광희동1가 구조, 성윤님 확정 2026-08-11)을 코드로 고정한다.
- 현장 CB 시트: ◆ 대수(기구물 시트 참조) → 모듈 행 → 합계 → 외함 → 요약 블록 → 요율 3종
- 현장 기구물 시트: ◆ 타입별 객실 수(노란 입력행) → 품목 행 → 총 수량표 → 요약 → 검산표
- 요율(확정 2026-08-11): 계약 실행가 = 실행 ×1.5 / 견적가 = 실행 ×1.8 / 설계예산가 = 계약가 ×1.6
- 단가는 「1.모듈총괄」 한 곳에서만 관리, 현장 시트는 INDEX/MATCH 참조
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MASTER_SHEET = "1.모듈총괄"
HISTORY_SHEET = "5.변경 이력"

# 색상 (v40 실측값)
YELLOW = "FFF2CC"   # 성윤님 입력칸 / 임시단가
ORANGE = "FBE5D6"   # 신규·미확정
GREEN = "C6E0B4"    # 실행가 행
GRAY = "D9D9D9"     # 외함 행
RED_FONT = "C00000"  # [확인] 계열

RATE_CONTRACT = 1.5   # 계약 실행가
RATE_QUOTE = 1.8      # 견적가 (업체 30% UP)
RATE_BUDGET = 1.6     # 설계예산가 = 계약 실행가 × 1.6

THIN = Border(*[Side(style="thin")] * 4)


def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def master_price_formula(name_cell):
    """총괄 참조 단가 수식. 가격 관리는 총괄 한 곳에서만 한다."""
    return (f"=IFERROR(INDEX('{MASTER_SHEET}'!$D:$D,"
            f"MATCH({name_cell},'{MASTER_SHEET}'!$B:$B,0)),\"\")")


def col_range(first_idx, count):
    """first_idx(1-base)부터 count개의 열 문자를 돌려준다."""
    return [get_column_letter(first_idx + i) for i in range(count)]


def set_cell(ws, addr, value, *, bold=False, fill_color=None, wrap=False,
             number_format=None, align=None, font_color=None, size=None):
    c = ws[addr]
    c.value = value
    if bold or font_color or size:
        c.font = Font(bold=bold, color=font_color, size=size or 10)
    if fill_color:
        c.fill = fill(fill_color)
    if wrap or align:
        c.alignment = Alignment(wrap_text=wrap, horizontal=align, vertical="center")
    if number_format:
        c.number_format = number_format
    return c


def find_master_names(wb):
    """총괄 B열의 모듈명 전수 → {이름: 행번호}. 명칭 매칭이 금액의 생명선이다."""
    ws = wb[MASTER_SHEET]
    names = {}
    for row in ws.iter_rows(min_col=2, max_col=2):
        v = row[0].value
        if isinstance(v, str) and v.strip() and row[0].row >= 5:
            names[v.strip()] = row[0].row
    return names


def last_data_row(ws, col=1, scan=2000):
    last = 0
    for r in ws.iter_rows(min_row=1, max_row=scan):
        if any(c.value is not None and str(c.value).strip() != "" for c in r):
            last = r[0].row
    return last
