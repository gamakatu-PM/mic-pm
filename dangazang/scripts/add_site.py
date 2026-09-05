# -*- coding: utf-8 -*-
"""현장 판독 JSON → CB모듈_단가장 현장 시트 생성/갱신 (표준 레이아웃 v70 SK하이닉스 세대).

코워크에서 손으로 하던 시트 조립을 코드로 고정한 것. 도면 판독(무엇이 몇 개인지)만
사람이/Claude가 JSON으로 만들면, 이후는 전부 결정적(deterministic)으로 처리된다:

  1. 현장 CB 시트 + 기구물 시트를 표준 레이아웃으로 생성 (기존 시트가 있으면 교체)
     - 시트 오른쪽에 「■ 요율 입력」 노란 칸(조립비율/계약/견적/예산)을 만들고
       모든 요율 계산이 그 칸을 참조한다. 기본값 0.3 / 1.5 / 1.8 / 2.1
       (성윤 차장 확정 2026-08-13 — 구 160% 폐기. 현장별 조정은 노란 칸에서)
  2. 총괄에 없는 모듈 → 「1.모듈총괄」 맨 끝에 주황 행 자동 등록 (가격은 성윤님 몫)
  3. 「0.요율판」이 있으면 이 현장 행을 자동 등록 (배수·총액이 요율판에 모임)
  4. 「5.변경 이력」에 A/B 등급 이력 자동 기입
  5. 출력 파일은 버전 +1 (원본은 절대 덮어쓰지 않는다)

사용:
    python3 add_site.py 단가장_v70.xlsx sites/현장명.json [-o out.xlsx]
"""
import argparse
import datetime
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from common import (GRAY, GREEN, HISTORY_SHEET, MASTER_SHEET, ORANGE,
                    RATE_ASSEMBLY, RATE_BUDGET, RATE_CONTRACT, RATE_QUOTE,
                    YELLOW, find_master_names, last_data_row,
                    master_price_formula, set_cell)

TODAY = datetime.date.today().isoformat()
NUMFMT = "#,##0"
RATE_PLATE = "0.요율판"


def qty_or_blank(v):
    return v if v is not None else None


def write_rate_block(ws, col, *, assembly=True):
    """시트 오른쪽 「■ 요율 입력」 블록. 반환: {'asm','c15','c18','c21'} → 절대참조 문자열.

    노란 칸이 배수의 단일 출처다. 라벨·계산 전부 이 칸을 참조하므로
    차장님이 칸 하나를 고치면 시트 전체 금액이 따라간다.
    """
    L = get_column_letter
    k, l, m = L(col), L(col + 1), L(col + 2)
    set_cell(ws, f"{k}3", "■ 요율 입력  — 노란 칸 숫자만 고치세요 (전부 실행 기준)", bold=True)
    for i, h in enumerate(["항목", "배수", "설명"]):
        set_cell(ws, f"{L(col+i)}4", h, bold=True, align="center")
    r = 5
    refs = {}
    rows = []
    if assembly:
        rows.append(("asm", "조립비율", RATE_ASSEMBLY,
                     "CB 내부 자재비 × 이 비율 (외함 제외)"))
    rows += [
        ("c15", "계약 실행가 배수", RATE_CONTRACT, "실행 × 이 배수 = 업체와 최종 계약할 금액"),
        ("c18", "견적가 배수", RATE_QUOTE, "실행 × 이 배수 = 업체에 보낼 견적 금액"),
        ("c21", "설계 예산가 배수", RATE_BUDGET,
         "★ 실행 × 이 배수 = 설계 예산가 (2026-08-13 확정, 구 160% 폐기)"),
    ]
    for key, label, val, desc in rows:
        set_cell(ws, f"{k}{r}", label)
        set_cell(ws, f"{l}{r}", val, fill_color=YELLOW, align="center")
        set_cell(ws, f"{m}{r}", desc)
        refs[key] = f"${l}${r}"
        r += 1
    set_cell(ws, f"{k}{r+1}",
             "※ 노란 칸 숫자만 고치면 시트 전체 계산이 전부 따라 바뀝니다.")
    set_cell(ws, f"{k}{r+2}",
             "※ 현장 규모·발주처에 따라 배수는 달라집니다. 정하는 것은 성윤 차장입니다.")
    return refs


# ---------------------------------------------------------------- CB 시트
def build_cb_sheet(wb, site, data, fixture_sheet_name):
    name = site
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    cb = data["cb"]
    types = cb["types"]
    T = len(types)
    tcols = list(range(3, 3 + T))                      # C..
    total_c, price_c = 3 + T, 4 + T                    # 총합, 단가
    money = list(range(5 + T, 5 + 2 * T))              # 금액 블록
    note_c = 5 + 2 * T
    rate_c = note_c + 2                                # 요율 블록 시작
    L = get_column_letter
    t0, t1 = L(tcols[0]), L(tcols[-1])

    set_cell(ws, "A1", cb.get("title") or f"현장: {site}", bold=True)
    if cb.get("note"):
        set_cell(ws, "A2", cb["note"])
    set_cell(ws, f"{t0}3", "수 량 (CB 1대당 · 타입별)", bold=True)
    set_cell(ws, f"{L(money[0])}3", "금 액 (CB 1대당 × 단가)", bold=True)
    headers = (["모듈명 (형번)", "도면 표기 (규격·원문)"] + types
               + ["총합\n(자동)", "실행단가\n(자동)"] + types + ["비고"])
    for i, h in enumerate(headers, start=1):
        set_cell(ws, f"{L(i)}4", h, bold=True, wrap=True, align="center")
    R = write_rate_block(ws, rate_c, assembly=True)

    # 5행 ◆ 대수 — 기구물 시트의 객실 수를 참조 (이중 관리 금지)
    set_cell(ws, "A5", "◆ CB 대수 [타입별 객실 수]", bold=True, fill_color=YELLOW)
    if fixture_sheet_name:
        set_cell(ws, "B5",
                 f"※ 「{fixture_sheet_name}」 시트 5행을 자동 참조합니다. 수정은 그 시트에서.")
        for i, c in enumerate(tcols):
            src = L(4 + i)
            set_cell(ws, f"{L(c)}5", f"='{fixture_sheet_name}'!{src}5",
                     fill_color=YELLOW, number_format=NUMFMT)
    else:
        set_cell(ws, "B5", "※ 타입별 대수를 노란칸에 입력하세요.")
        for c in tcols:
            set_cell(ws, f"{L(c)}5", None, fill_color=YELLOW, number_format=NUMFMT)
    set_cell(ws, f"{L(total_c)}5", f"=SUM({t0}5:{t1}5)", bold=True, number_format=NUMFMT)
    if data.get("fixtures_sheet", {}).get("rooms_note"):
        set_cell(ws, f"{L(note_c)}5", data["fixtures_sheet"]["rooms_note"])

    def item_row(r, item):
        set_cell(ws, f"A{r}", item["name"])
        set_cell(ws, f"B{r}", item.get("drawing"))
        for t, c in zip(types, tcols):
            set_cell(ws, f"{L(c)}{r}", qty_or_blank(item["qty"].get(t)),
                     number_format=NUMFMT)
        set_cell(ws, f"{L(total_c)}{r}",
                 f'=IF(SUM(${t0}$5:${t1}$5)=0,"",SUMPRODUCT(${t0}$5:${t1}$5,{t0}{r}:{t1}{r}))',
                 number_format=NUMFMT)
        pcol = L(price_c)
        if item.get("unit_price_value") is not None:
            # 총괄 미등록 상태에서 성윤님이 지정한 값 — 노란칸으로 두고 비고에 남긴다
            set_cell(ws, f"{pcol}{r}", item["unit_price_value"],
                     fill_color=YELLOW, number_format=NUMFMT)
        else:
            set_cell(ws, f"{pcol}{r}", master_price_formula(f"A{r}"),
                     number_format=NUMFMT)
        for i, c in enumerate(money):
            q = f"{L(tcols[i])}{r}"
            set_cell(ws, f"{L(c)}{r}",
                     f'=IF(OR({q}="",${pcol}{r}=""),"",{q}*${pcol}{r})',
                     number_format=NUMFMT)
        set_cell(ws, f"{L(note_c)}{r}", item.get("note"))

    r = 6
    first_mod = r
    for item in cb["modules"]:
        item_row(r, item)
        r += 1
    last_mod = r - 1
    sum_row = r
    set_cell(ws, f"A{sum_row}", "합계", bold=True)
    for i, c in enumerate(money):
        terms = "+".join(
            f"IFERROR({L(tcols[i])}{k}*${L(price_c)}{k},0)"
            for k in range(first_mod, last_mod + 1))
        set_cell(ws, f"{L(c)}{sum_row}", f"={terms}", bold=True, number_format=NUMFMT)
    set_cell(ws, f"{L(note_c)}{sum_row}", "CB 1대당 자재비 소계 (외함 제외)")
    r += 2

    enc_rows = []
    for item in cb.get("enclosures", []):
        item_row(r, item)
        for col in range(1, note_c + 1):
            ws.cell(row=r, column=col).fill = openpyxl.styles.PatternFill(
                start_color=GRAY, end_color=GRAY, fill_type="solid")
        enc_rows.append(r)
        r += 1
    r += 1

    # ---- 요약 블록 — 배수는 전부 「■ 요율 입력」 칸 참조 (하드코딩 금지)
    def summary(label, note, fml, *, fill_color=None, per_type=True):
        nonlocal r
        set_cell(ws, f"A{r}", label, bold=True, fill_color=fill_color)
        cols = money if per_type else [money[0]]
        for i, c in enumerate(cols):
            set_cell(ws, f"{L(c)}{r}", fml(i), fill_color=fill_color,
                     number_format=NUMFMT)
        set_cell(ws, f"{L(note_c)}{r}", note)
        row = r
        r += 1
        return row

    m0 = lambda i: L(money[i])
    q0 = lambda i: L(tcols[i])
    r_mat = summary("▣ 자재비 소계 (CB 1대당, 외함 제외)", "외함 제외 자재비",
                    lambda i: f"={m0(i)}{sum_row}")
    r_asm = summary(f'=CONCATENATE("▣ 조립비 (자재 ×",TEXT({R["asm"]},"0%"),")")',
                    "조립비 = 자재비 × 요율칸(조립비율)",
                    lambda i: f"=ROUND({m0(i)}{r_mat}*{R['asm']},0)")
    r_unit = summary("■ CB 1대당 실행가 (견적 CONTROL BOX 단가)",
                     "자재비 + 조립비  ← 견적서 CONTROL BOX 라인",
                     lambda i: f"={m0(i)}{r_mat}+{m0(i)}{r_asm}", fill_color=GREEN)
    if enc_rows:
        enc_terms = lambda i: "+".join(
            f"IFERROR({q0(i)}{er}*${L(price_c)}{er},0)" for er in enc_rows)
        r_enc = summary("■ CB외함 실행가 (견적 별도 라인)",
                        "외함 단가  ← 견적서 CB외함 라인",
                        lambda i: f"={enc_terms(i)}", fill_color=GRAY)
    else:
        r_enc = summary("■ CB외함 실행가 (견적 별도 라인)", "외함 없음",
                        lambda i: "=0", fill_color=GRAY)
    r_exec = summary("▣ 실행 합계 (본체+외함)", "본체 + 외함",
                     lambda i: f"={m0(i)}{r_unit}+{m0(i)}{r_enc}")
    r += 1
    r_exec_tot = summary("▣ 실행 합계 (대수 반영)", "1대당 실행 × 타입별 대수",
                         lambda i: f'=IF({q0(i)}$5="","",{m0(i)}{r_exec}*{q0(i)}$5)')
    r_c15 = summary(f'=CONCATENATE("▷ 계약 실행가 (실행 ×",TEXT({R["c15"]},"0.0"),")")',
                    "계약 실행가 — 배수는 요율칸에서",
                    lambda i: f"=ROUND({m0(i)}{r_exec}*{R['c15']},0)")
    r_c18 = summary(f'=CONCATENATE("▷ 견적가 (실행 ×",TEXT({R["c18"]},"0.0"),")")',
                    "견적가 — 배수는 요율칸에서",
                    lambda i: f"=ROUND({m0(i)}{r_exec}*{R['c18']},0)")
    r_c21 = summary(f'=CONCATENATE("▷ 설계 예산가 (실행 ×",TEXT({R["c21"]},"0.0"),")")',
                    "설계 예산가 — 배수는 요율칸에서 (구 160% 폐기)",
                    lambda i: f"=ROUND({m0(i)}{r_exec}*{R['c21']},0)")
    r_q_tot = summary(f'=CONCATENATE("▷ 견적 합계 (계약 ×",TEXT({R["c15"]},"0.0"),", 대수 반영)")',
                      "계약 실행가 기준",
                      lambda i: f'=IF({q0(i)}$5="","",{m0(i)}{r_c15}*{q0(i)}$5)')
    r += 1
    m_first, m_last = L(money[0]), L(money[-1])
    st_exec = summary("★ 현장 실행 총액", "",
                      lambda i: f"=SUM({m_first}{r_exec_tot}:{m_last}{r_exec_tot})",
                      per_type=False)
    st_c15 = summary(f'=CONCATENATE("★ 현장 계약 실행가 총액 (실행 ×",TEXT({R["c15"]},"0.0"),")")',
                     "", lambda i: f"=SUM({m_first}{r_q_tot}:{m_last}{r_q_tot})",
                     per_type=False)
    st_c18 = summary(f'=CONCATENATE("★ 현장 견적가 총액 (실행 ×",TEXT({R["c18"]},"0.0"),")")',
                     "", lambda i: f"=SUMPRODUCT({m_first}{r_c18}:{m_last}{r_c18},${t0}$5:${t1}$5)",
                     per_type=False)
    summary(f'=CONCATENATE("★ 현장 설계 예산가 총액 (실행 ×",TEXT({R["c21"]},"0.0"),")")',
            "", lambda i: f"=SUMPRODUCT({m_first}{r_c21}:{m_last}{r_c21},${t0}$5:${t1}$5)",
            per_type=False)
    r += 1
    for fn in cb.get("footnotes", []):
        set_cell(ws, f"A{r}", fn)
        r += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 40
    for c in tcols + [total_c, price_c] + money:
        ws.column_dimensions[L(c)].width = 11
    ws.column_dimensions[L(note_c)].width = 44
    for c in (rate_c, rate_c + 1, rate_c + 2):
        ws.column_dimensions[L(c)].width = 16
    # 요율판 등록용 좌표
    ws._km = {"rate": R, "exec_total": f"{m_first}{st_exec}",
              "quote_total": f"{m_first}{st_c18}", "size": f"{L(total_c)}5",
              "asm": R.get("asm")}
    return ws


# ---------------------------------------------------------------- 기구물 시트
def build_fixture_sheet(wb, site, data):
    fx = data["fixtures_sheet"]
    name = f"{site} 기구물"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    types = fx["types"]
    T = len(types)
    L = get_column_letter
    tcols = list(range(4, 4 + T))                      # D.. (C=중앙장비)
    total_c, price_c = 4 + T, 5 + T                    # 객실 총합, 단가
    money_room, money_ctr = 6 + T, 7 + T
    note_c = 8 + T
    rate_c = note_c + 2
    t0, t1 = L(tcols[0]), L(tcols[-1])
    total_rooms = sum(v for v in fx["rooms_by_type"].values() if v) or ""

    set_cell(ws, "A1", fx.get("title") or f"현장: {site} — 객실 기구물 + 중앙장비", bold=True)
    if fx.get("note"):
        set_cell(ws, "A2", fx["note"])
    set_cell(ws, f"{t0}3", "수 량 (1실당 · 타입별)", bold=True)
    set_cell(ws, f"{L(total_c)}3", "수 량 (총합·자동)", bold=True)
    set_cell(ws, f"{L(money_room)}3", "합 계 (금액)", bold=True)
    headers = (["모듈명 (형번)", "도면 표기 (규격·원문)", "중앙장비\n(1식)"] + types
               + [f"객실 ({total_rooms}실)", "실행단가\n(자동)",
                  f"객실 ({total_rooms}실)", "중앙장비\n(1식)", "비고"])
    for i, h in enumerate(headers, start=1):
        set_cell(ws, f"{L(i)}4", h, bold=True, wrap=True, align="center")
    R = write_rate_block(ws, rate_c, assembly=False)  # 기구물엔 조립비 없음

    set_cell(ws, "A5", "◆ 타입별 객실 수  (평면도 기준 · 노란칸 입력)", bold=True,
             fill_color=YELLOW)
    set_cell(ws, "B5", f"타입별 몇 실인지. 합계가 {total_rooms}이 되어야 합니다.")
    set_cell(ws, "C5", 1, number_format=NUMFMT)
    for t, c in zip(types, tcols):
        set_cell(ws, f"{L(c)}5", fx["rooms_by_type"].get(t), fill_color=YELLOW,
                 number_format=NUMFMT)
    set_cell(ws, f"{L(total_c)}5", f"=SUM({t0}5:{t1}5)", bold=True, number_format=NUMFMT)
    set_cell(ws, f"{L(note_c)}5", fx.get("rooms_note"))

    r = 6
    first = r
    for item in fx["fixtures"]:
        set_cell(ws, f"A{r}", item["name"])
        set_cell(ws, f"B{r}", item.get("drawing"))
        set_cell(ws, f"C{r}", qty_or_blank(item.get("central")), number_format=NUMFMT)
        for t, c in zip(types, tcols):
            set_cell(ws, f"{L(c)}{r}", qty_or_blank(item["qty"].get(t)),
                     number_format=NUMFMT)
        set_cell(ws, f"{L(total_c)}{r}",
                 f'=IF(SUM(${t0}$5:${t1}$5)=0,"",SUMPRODUCT(${t0}$5:${t1}$5,{t0}{r}:{t1}{r}))',
                 number_format=NUMFMT)
        pcol = L(price_c)
        if item.get("unit_price_value") is not None:
            set_cell(ws, f"{pcol}{r}", item["unit_price_value"], fill_color=YELLOW,
                     number_format=NUMFMT)
        else:
            set_cell(ws, f"{pcol}{r}", master_price_formula(f"A{r}"),
                     number_format=NUMFMT)
        set_cell(ws, f"{L(money_room)}{r}",
                 f'=IF(OR({L(total_c)}{r}="",${pcol}{r}=""),"",{L(total_c)}{r}*${pcol}{r})',
                 number_format=NUMFMT)
        set_cell(ws, f"{L(money_ctr)}{r}",
                 f'=IF(OR(C{r}="",${pcol}{r}=""),"",C{r}*${pcol}{r})',
                 number_format=NUMFMT)
        set_cell(ws, f"{L(note_c)}{r}", item.get("note"))
        r += 1
    last = r - 1
    sum_row = r
    set_cell(ws, f"A{sum_row}", "합계", bold=True)
    for colx, qcol in ((money_room, total_c), (money_ctr, 3)):
        terms = "+".join(f"IFERROR({L(qcol)}{k}*${L(price_c)}{k},0)"
                         for k in range(first, last + 1))
        set_cell(ws, f"{L(colx)}{sum_row}", f"={terms}", bold=True, number_format=NUMFMT)
    set_cell(ws, f"{L(note_c)}{sum_row}", "객실 총액 / 중앙장비 총액")
    r += 2

    # ---- ★ 객실 타입별 총 수량표 (전부 자동)
    set_cell(ws, f"A{r}", "★ 객실 타입별 총 수량표  (= 위 표 1실당 수량 × 타입별 객실 수 — 전부 자동)",
             bold=True)
    r += 1
    for i, h in enumerate(headers, start=1):
        set_cell(ws, f"{L(i)}{r}", h, bold=True, wrap=True, align="center")
    r += 1
    tq_first = r
    for k in range(first, last + 1):
        set_cell(ws, f"A{r}", f"=A{k}")
        set_cell(ws, f"B{r}", f"=B{k}")
        set_cell(ws, f"C{r}", f'=IF(C{k}="","",C{k}*$C$5)', number_format=NUMFMT)
        for c in tcols:
            cl = L(c)
            set_cell(ws, f"{cl}{r}",
                     f'=IF(OR({cl}{k}="",{cl}$5=""),"",{cl}{k}*{cl}$5)',
                     number_format=NUMFMT)
        for c in (total_c, price_c, money_room, money_ctr):
            cl = L(c)
            set_cell(ws, f"{cl}{r}", f'=IF({cl}{k}="","",{cl}{k})', number_format=NUMFMT)
        set_cell(ws, f"{L(note_c)}{r}", fx["fixtures"][k - first].get("note"))
        r += 1
    tq_last = r - 1
    set_cell(ws, f"A{r}", "합계", bold=True)
    for c in [3] + tcols + [total_c]:
        cl = L(c)
        set_cell(ws, f"{cl}{r}",
                 f'=IF(SUM({cl}{tq_first}:{cl}{tq_last})=0,"",SUM({cl}{tq_first}:{cl}{tq_last}))',
                 bold=True, number_format=NUMFMT)
    for c in (money_room, money_ctr):
        # 금액 열은 일반 SUM — ""를 흘려보내면 아래 요율 행이 #VALUE!로 죽는다
        cl = L(c)
        set_cell(ws, f"{cl}{r}", f"=SUM({cl}{tq_first}:{cl}{tq_last})",
                 bold=True, number_format=NUMFMT)
    set_cell(ws, f"{L(note_c)}{r}", "객실 총액 / 중앙장비 총액")
    tq_sum = r
    r += 2

    mr, mc = L(money_room), L(money_ctr)
    def srow(label, froom, fctr, note, gap=0):
        nonlocal r
        set_cell(ws, f"A{r}", label, bold=True)
        set_cell(ws, f"{mr}{r}", froom, number_format=NUMFMT)
        if fctr:
            set_cell(ws, f"{mc}{r}", fctr, number_format=NUMFMT)
        set_cell(ws, f"{L(note_c)}{r}", note)
        row = r
        r += 1 + gap
        return row

    r_mat = srow("▣ 자재비 총액 (객실 / 중앙장비)", f"={mr}{tq_sum}", f"={mc}{tq_sum}",
                 "기구물에는 조립비를 붙이지 않습니다")
    r_exec = srow("▣ 실행 합계", f"={mr}{r_mat}", f"={mc}{r_mat}",
                  "자재비 = 실행 (조립비 없음)", gap=1)
    r15 = srow(f'=CONCATENATE("▷ 계약 실행가 총액 (실행 ×",TEXT({R["c15"]},"0.0"),")")',
               f"=ROUND({mr}{r_exec}*{R['c15']},0)",
               f"=ROUND({mc}{r_exec}*{R['c15']},0)",
               "계약 실행가 — 배수는 요율칸에서")
    r18 = srow(f'=CONCATENATE("▷ 견적가 총액 (실행 ×",TEXT({R["c18"]},"0.0"),")")',
               f"=ROUND({mr}{r_exec}*{R['c18']},0)",
               f"=ROUND({mc}{r_exec}*{R['c18']},0)",
               "견적가 — 배수는 요율칸에서")
    r21 = srow(f'=CONCATENATE("▷ 설계 예산가 총액 (실행 ×",TEXT({R["c21"]},"0.0"),")")',
               f"=ROUND({mr}{r_exec}*{R['c21']},0)",
               f"=ROUND({mc}{r_exec}*{R['c21']},0)",
               "설계 예산가 — 배수는 요율칸에서 (구 160% 폐기)", gap=1)
    st_exec = srow("★ 현장 실행 총액", f"={mr}{r_exec}+{mc}{r_exec}", None, "")
    srow(f'=CONCATENATE("★ 현장 계약 실행가 총액 (실행 ×",TEXT({R["c15"]},"0.0"),")")',
         f"={mr}{r15}+{mc}{r15}", None, "")
    st_c18 = srow(f'=CONCATENATE("★ 현장 견적가 총액 (실행 ×",TEXT({R["c18"]},"0.0"),")")',
                  f"={mr}{r18}+{mc}{r18}", None, "")
    srow(f'=CONCATENATE("★ 현장 설계 예산가 총액 (실행 ×",TEXT({R["c21"]},"0.0"),")")',
         f"={mr}{r21}+{mc}{r21}", None, "", gap=2)

    # ---- 검산표 — 계통도 수량표와 자동 대조
    if fx.get("checks"):
        set_cell(ws, f"A{r}", "■ 검산 — 계통도 수량표와 대조 (자동)", bold=True)
        r += 1
        for i, h in enumerate(["품목", "계통도 수량표", "시트 총합", "판정"], start=1):
            set_cell(ws, f"{L(i)}{r}", h, bold=True)
        r += 1
        for chk in fx["checks"]:
            set_cell(ws, f"A{r}", chk["item"])
            set_cell(ws, f"B{r}", chk["expected"], number_format=NUMFMT)
            set_cell(ws, f"C{r}",
                     f'=IFERROR(INDEX(${L(total_c)}${tq_first}:${L(total_c)}${tq_last},'
                     f'MATCH(A{r},$A${tq_first}:$A${tq_last},0)),"")',
                     number_format=NUMFMT)
            set_cell(ws, f"D{r}", f'=IF(C{r}="","-",IF(C{r}=B{r},"OK","불일치"))')
            r += 1
        r += 1
    for fn in fx.get("footnotes", []):
        set_cell(ws, f"A{r}", fn)
        r += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 40
    for c in [3] + tcols + [total_c, price_c, money_room, money_ctr]:
        ws.column_dimensions[L(c)].width = 11
    ws.column_dimensions[L(note_c)].width = 44
    for c in (rate_c, rate_c + 1, rate_c + 2):
        ws.column_dimensions[L(c)].width = 16
    ws._km = {"rate": R, "exec_total": f"{mr}{st_exec}",
              "quote_total": f"{mr}{st_c18}", "size": f"{L(total_c)}5",
              "asm": None}
    return ws


# ---------------------------------------------------------------- 총괄·요율판·이력
def register_new_modules(wb, data):
    """총괄에 없는 모듈명 → 맨 끝 주황 행으로 등록. 가격은 성윤님이 채운다."""
    names = find_master_names(wb)
    items = list(data["cb"]["modules"]) + list(data["cb"].get("enclosures", []))
    if "fixtures_sheet" in data:
        items += data["fixtures_sheet"]["fixtures"]
    ws = wb[MASTER_SHEET]
    row = last_data_row(ws) + 1
    added = []
    for item in items:
        nm = item["name"].strip()
        if nm in names:
            continue
        gubun = "외함" if "외함" in nm else "신규"
        set_cell(ws, f"A{row}", gubun)
        set_cell(ws, f"B{row}", nm)
        set_cell(ws, f"C{row}", (item.get("drawing") or "")[:60])
        note = "[신규·단가확인] " + (data["site"])
        if item.get("unit_price_value") is not None:
            note += f" / 현장 시트에 임시값 {item['unit_price_value']:,.0f} 있음"
        set_cell(ws, f"E{row}", note)
        for col in "ABCDE":
            ws[f"{col}{row}"].fill = openpyxl.styles.PatternFill(
                start_color=ORANGE, end_color=ORANGE, fill_type="solid")
        added.append(nm)
        names[nm] = row
        row += 1
    return added


def register_rate_plate(wb, sheets):
    """「0.요율판」에 이 현장 행 등록 — 배수·총액을 한 곳에서 보는 판."""
    if RATE_PLATE not in wb.sheetnames:
        return []
    ws = wb[RATE_PLATE]
    existing = {str(ws.cell(row=r, column=1).value).strip()
                for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=1).value}
    row = last_data_row(ws) + 1
    added = []
    for s in sheets:
        if s.title in existing:
            # 기존 행은 그대로 두고 새 행을 추가하지 않는다 (중복 방지) — 참조 갱신은 수동 확인
            continue
        km = s._km
        set_cell(ws, f"A{row}", s.title)
        set_cell(ws, f"B{row}", f"='{s.title}'!{km['size']}")
        set_cell(ws, f"C{row}", f"='{s.title}'!{km['asm'].replace('$','')}" if km["asm"] else "—")
        set_cell(ws, f"D{row}", f"='{s.title}'!{km['rate']['c15'].replace('$','')}")
        set_cell(ws, f"E{row}", f"='{s.title}'!{km['rate']['c18'].replace('$','')}")
        set_cell(ws, f"F{row}", f"='{s.title}'!{km['rate']['c21'].replace('$','')}")
        set_cell(ws, f"G{row}", f"='{s.title}'!{km['exec_total']}")
        set_cell(ws, f"H{row}", f"='{s.title}'!{km['quote_total']}")
        added.append(s.title)
        row += 1
    return added


def append_history(wb, version, entries):
    ws = wb[HISTORY_SHEET]
    row = last_data_row(ws) + 1
    for grade, target, what, before, revert in entries:
        vals = [f"v{version}", TODAY, grade, target, what, before, revert]
        for i, v in enumerate(vals, start=1):
            set_cell(ws, f"{get_column_letter(i)}{row}", v)
        row += 1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("site_json")
    ap.add_argument("-o", "--output")
    args = ap.parse_args()

    data = json.load(open(args.site_json, encoding="utf-8"))
    site = data["site"]
    wb = openpyxl.load_workbook(args.workbook)

    m = re.search(r"_v(\d+)", Path(args.workbook).name)
    version = (int(m.group(1)) if m else 0) + 1
    out = args.output or str(Path(args.workbook).with_name(
        f"CB모듈_단가장_v{version}.xlsx"))

    replaced_cb = site in wb.sheetnames
    fixture_name = f"{site} 기구물" if "fixtures_sheet" in data else None
    replaced_fx = fixture_name in wb.sheetnames if fixture_name else False

    sheets = []
    if fixture_name:
        sheets.append(build_fixture_sheet(wb, site, data))
    sheets.append(build_cb_sheet(wb, site, data, fixture_name))
    added = register_new_modules(wb, data)
    plate_added = register_rate_plate(wb, sheets)

    entries = []
    grade_cb = "B" if replaced_cb else "A"
    entries.append((grade_cb, f"{site} 시트",
                    f"add_site.py 자동 생성 (표준 레이아웃·요율칸 1.5/1.8/2.1, 입력: {Path(args.site_json).name})",
                    "기존 시트 전체" if replaced_cb else "(없던 시트)",
                    f"v{version-1} 파일에 원래 시트 그대로 있음" if replaced_cb else "시트 삭제"))
    if fixture_name:
        grade_fx = "B" if replaced_fx else "A"
        entries.append((grade_fx, f"{fixture_name} 시트",
                        "add_site.py 자동 생성 (표준 레이아웃·요율칸)",
                        "기존 시트 전체" if replaced_fx else "(없던 시트)",
                        f"v{version-1} 파일에 원래 시트 그대로 있음" if replaced_fx else "시트 삭제"))
    if added:
        entries.append(("A", MASTER_SHEET,
                        f"신규 모듈 {len(added)}건 주황 등록: " + ", ".join(added),
                        "(없던 행)", "해당 행 삭제"))
    if plate_added:
        entries.append(("A", RATE_PLATE,
                        f"현장 행 {len(plate_added)}건 등록: " + ", ".join(plate_added),
                        "(없던 행)", "해당 행 삭제"))
    append_history(wb, version, entries)

    wb.save(out)
    print(json.dumps({
        "output": out, "version": version, "site": site,
        "cb_sheet": "replaced" if replaced_cb else "added",
        "fixture_sheet": ("replaced" if replaced_fx else "added") if fixture_name else None,
        "new_modules_registered": added,
        "rate_plate_registered": plate_added,
        "rates_default": {"assembly": RATE_ASSEMBLY, "contract": RATE_CONTRACT,
                          "quote": RATE_QUOTE, "budget": RATE_BUDGET},
        "history_rows": len(entries),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
