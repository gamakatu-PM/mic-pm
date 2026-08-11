# -*- coding: utf-8 -*-
"""기존 CB모듈_단가장에서 현장 시트를 JSON으로 역추출한다.

도면 판독 결과(무엇이 몇 개인지)는 사람이/Claude가 만들지만, 일단 시트에 들어간
판독 결과는 이 스크립트로 데이터화해서 git으로 이력 관리한다.

사용:
    python3 extract_site.py 단가장.xlsx "광희동1가" > sites/광희동1가.json
    (기구물 시트 "광희동1가 기구물"이 있으면 자동으로 함께 추출)
"""
import json
import re
import sys

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula


def cellv(c):
    v = c.value
    if isinstance(v, ArrayFormula):
        return v.text
    return v


def is_formula(c):
    v = cellv(c)
    return isinstance(v, str) and v.startswith("=")


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, str):
        return None
    return int(v) if float(v) == int(v) else float(v)


def header_types(ws, header_row=4, first_col=3):
    """4행 헤더에서 타입 열 목록을 뽑는다. '총합'을 만나면 끝."""
    types, cols = [], []
    for col in range(first_col, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v is None:
            break
        s = str(v).replace("\n", " ")
        if "총합" in s or "객실" in s and "실)" in s:
            break
        types.append(s.strip())
        cols.append(col)
    return types, cols


def extract_cb_sheet(ws):
    types, tcols = header_types(ws)
    ncols = len(tcols)
    qty_total_col = tcols[-1] + 1
    price_col = qty_total_col + 1
    note_col = price_col + 1 + ncols

    out = {
        "title": ws["A1"].value,
        "note": ws["A2"].value,
        "types": types,
        "modules": [],
        "enclosures": [],
        "footnotes": [],
    }
    section = "modules"
    for r in range(6, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a is None:
            continue
        a = str(a)
        if a.startswith("합계"):
            section = "enclosures"
            continue
        if a.startswith(("▣", "■", "◆", "▷", "★")):
            break
        if a.startswith("※"):
            out["footnotes"].append(a)
            continue
        item = {
            "name": a.strip(),
            "drawing": ws.cell(row=r, column=2).value,
            "qty": {t: num(ws.cell(row=r, column=c).value)
                    for t, c in zip(types, tcols)},
            "note": ws.cell(row=r, column=note_col).value,
        }
        pc = ws.cell(row=r, column=price_col)
        if not is_formula(pc) and pc.value is not None:
            item["unit_price_value"] = num(pc.value)  # 값으로 굳은 단가 — 총괄 미등록 신호
        target = "enclosures" if "외함" in a else section
        out[target].append(item)
    # 요약·footnote 구간의 ※ 행도 수집
    for r in range(6, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.startswith("※") and a not in out["footnotes"]:
            out["footnotes"].append(a)
    return out


def extract_fixture_sheet(ws):
    # 기구물: C=중앙장비(1식), D..=타입, 그다음 객실총합/단가/객실금액/중앙금액/비고
    types, tcols = header_types(ws, first_col=4)
    out = {
        "title": ws["A1"].value,
        "note": ws["A2"].value,
        "types": types,
        "rooms_by_type": {},
        "rooms_note": None,
        "fixtures": [],
        "checks": [],
        "footnotes": [],
    }
    note_col = tcols[-1] + 5
    # 5행 = ◆ 타입별 객실 수
    for t, c in zip(types, tcols):
        out["rooms_by_type"][t] = num(ws.cell(row=5, column=c).value)
    out["rooms_note"] = ws.cell(row=5, column=note_col).value
    for r in range(6, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a is None:
            continue
        a = str(a)
        if a.startswith("합계") or a.startswith("★"):
            break
        item = {
            "name": a.strip(),
            "drawing": ws.cell(row=r, column=2).value,
            "central": num(ws.cell(row=r, column=3).value),
            "qty": {t: num(ws.cell(row=r, column=c).value)
                    for t, c in zip(types, tcols)},
            "note": ws.cell(row=r, column=note_col).value,
        }
        pc = ws.cell(row=r, column=tcols[-1] + 2)
        if not is_formula(pc) and pc.value is not None:
            item["unit_price_value"] = num(pc.value)
        out["fixtures"].append(item)
    # 검산표: "■ 검산" 이후 품목|계통도수량 행
    in_check = False
    for r in range(6, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.startswith("■ 검산"):
            in_check = True
            continue
        if in_check:
            if a is None or str(a).startswith(("※", "품목")):
                if isinstance(a, str) and a.startswith("※"):
                    break
                continue
            expected = num(ws.cell(row=r, column=2).value)
            if expected is not None:
                out["checks"].append({"item": str(a).strip(), "expected": expected})
        if isinstance(a, str) and a.startswith("※"):
            out["footnotes"].append(a)
    return out


def main():
    xlsx, site = sys.argv[1], sys.argv[2]
    wb = openpyxl.load_workbook(xlsx)
    data = {"site": site, "source": xlsx.split("/")[-1]}
    data["cb"] = extract_cb_sheet(wb[site])
    fx = f"{site} 기구물"
    if fx in wb.sheetnames:
        data["fixtures_sheet"] = extract_fixture_sheet(wb[fx])
    m = re.search(r"_v(\d+)", xlsx)
    if m:
        data["source_version"] = int(m.group(1))
    json.dump(data, sys.stdout, ensure_ascii=False, indent=2)
    print()


if __name__ == "__main__":
    main()
